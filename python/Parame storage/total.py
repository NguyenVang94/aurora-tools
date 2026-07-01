# =================================================================================
# IMPORT THƯ VIỆN 
# =================================================================================
import os
import sys
import shutil
import json
import glob
import pandas as pd
from datetime import datetime
import logging
import openpyxl
from charset_normalizer import detect
from colorama import init, Fore, Back, Style
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import TimeoutException
import gspread 
from google.oauth2.service_account import Credentials 


# =================================================================================
#       HẰNG SỐ VÀ CẤU HÌNH CHUNG
# =================================================================================

base_dir = os.getcwd() 

# HẰNG SỐ TW (Twitter)
TW_SOURCE_PATTERN = "入稿"
TW_SOURCE_EXTENSION = "*.xlsm"
TW_CSV_PATTERN = "掲載内容"
TW_CSV_EXTENSION = "*.csv"
TW_MEDIA_ID = "1021"
TW_ACTION = "add"
TW_MASTER_FILE = "Master.xlsx"
TW_SHEET_NAME = "ツイート設計シート"
TW_TWEET_ID_COL = "ツイートID"
TW_URL_COL = "ウェブサイトのURL"
TW_HEADER_ROW = 16  # Row 17 in Excel (0-based index)
TW_REQUIRED_CSV_COLS = [
    "ツイートID", "アカウントID", "キャンペーンID", "広告キャンペーン名",
    "広告グループID", "広告グループ名", "relationID"
]
TW_OUTPUT_COLUMNS = [
    "Action", "媒体ID", "CID", "アカウントID", "キャンペーンID", "キャンペーン名",
    "広告グループID", "広告グループ名", "広告ID", "告名", "relation ID",
    "キーワードID", "キーワード名", "マッチタイプ", "パラメ発行済みURL", "ドラフト停止日"
]

# HẰNG SỐ LINE (Line)
LINE_INPUT_DIR = os.path.join(base_dir, 'input')
LINE_MASTER_FILE = os.path.join(base_dir, 'master', 'Master.xlsx')
LINE_OUTPUT_DIR = os.path.join(base_dir,'output')
LINE_ERROR_LOG_FILE = os.path.join(LINE_OUTPUT_DIR, 'master2_error_log.txt')
LINE_DUPLICATE_LOG_FILE = os.path.join(LINE_OUTPUT_DIR, 'duplicate_campaigns.csv')
LINE_MISSING_COMBINATIONS_FILE = os.path.join(LINE_OUTPUT_DIR, 'missing_combinations.csv')
LINE_NAN_ROWS_FILE = os.path.join(LINE_OUTPUT_DIR, 'nan_rows_master2.csv')
LINE_MISSING_IDS_FILE = os.path.join(LINE_OUTPUT_DIR, 'missingLineID_Template2.csv')
LINE_MASTER_REQUIRED_COLS = ['アカウントID', 'CID']
LINE_MASTER2_REQUIRED_COLS = ['Campaign name', 'Campaign ID', 'Ad group name', 'Ad group ID', 'Ad name', 'Ad ID']
LINE_INPUT_REQUIRED_COLS = ['アカウントID', 'キャンペーン名', '広告グループ名', '広告名', 'リンク先URL1']
LINE_OUTPUT_COLUMNS = [
    'Action', '媒体ID', 'CID', 'アカウントID', 'キャンペーンID', 'キャンペーン名', '広告グループID',
    '広告グループ名', '広告ID', '告名', 'relation ID', 'キーワードID', 'キーワード名', 'マッチタイプ',
    'パラメ発行済みURL', 'ドラフト停止日'
]
LINE_MEDIA_ID = '9191'
LINE_ACTION = 'add'


# HẰNG SỐ FACEBOOK ---
FB_CSV_PATTERN = "export"
FB_CSV_EXTENSION = "*.csv"
FB_MEDIA_ID = "1020"
FB_ACTION = "add"
FB_OUTPUT_COLUMNS = [ 
    "Action", "媒体ID", "CID", "アカウントID", "キャンペーンID", "キャンペーン名",
    "広告グループID", "広告グループ名", "広告ID", "告名", "relation ID",
    "キーワードID", "キーワード名", "マッチタイプ", "パラメ発行済みURL", "ドラフト停止日"
]


# User
master_folder = os.path.join(base_dir, 'master')
file_name = 'User.xlsx'
user_file = os.path.join(master_folder, file_name)
username = ""
password = ""

URL_TARGET = "https://ca-rpa.cloud/en/parameter-storage/manager"
URL_LOGIN = "https://ca-rpa.cloud/login" 
URL_BASE = "https://ca-rpa.cloud/"
COOKIE_FILENAME = "cloud_cookies.json"
LOGIN_TIMEOUT = 300  
EXIT_DELAY = 10      
#  script_dir = os.path.dirname(os.path.abspath(__file__))
cookie_path = os.path.join(base_dir, COOKIE_FILENAME)


# Template2
TEMPLATE2_FILE = os.path.join(base_dir, 'master', 'Template2.xlsx')

# DATA BASE
UPLOAD_LOG_FILE = os.path.join(base_dir, 'master', 'upload_log.xlsx')

# OUTPUT
source_folder = os.path.join(base_dir, 'output')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.StreamHandler()]
)

# --- CÁC HÀM XỬ LÝ CHO TWITTER ---

def tw_setup_directories(source_dir, master_dir, output_dir):
    """Check and create output directory if needed for Twitter."""
    for directory in [source_dir, master_dir]:
        if not os.path.exists(directory):
            logging.error(f"フォルダが見つかりません。: {directory}")
            raise FileNotFoundError(f"フォルダが見つかりません。: {directory}")
    os.makedirs(output_dir, exist_ok=True)

def tw_load_master_file(master_path):
    """Read Master.xlsx and create a mapping dictionary for アカウントID -> CID for Twitter."""
    master_file = os.path.join(master_path, TW_MASTER_FILE)
    if not os.path.exists(master_file):
        logging.warning(f"{master_file}にてファイル「Master.xlsx」が見つかりません。")
        return {}
    try:
        df = pd.read_excel(master_file, engine='openpyxl', usecols=["アカウントID", "CID"])
        cid_map = {
            str(row["アカウントID"]).strip(): str(row["CID"]).strip()
            for _, row in df.iterrows()
            if not pd.isna(row["アカウントID"]) and not pd.isna(row["CID"])
        }
        logging.info(f"「Master.xlsx」にて、{len(cid_map)} 件のアカウントID -> CID の紐付けを検出しました。")
        return cid_map
    except FileNotFoundError:
        logging.error(f"「Master.xlsx」ファイルが存在しません:{master_file}")
        return {}
    except Exception as e:
        logging.error(f"「Master.xlsx」の読み込みエラー:{e}")
        return {}

# def tw_find_columns(df, target_tweet_id=TW_TWEET_ID_COL, target_url=TW_URL_COL, target_card_name="カード名"):
#     """Dynamically find columns in DataFrame that match or contain target column names, ignoring case and spaces."""
#     found_tweet_id, found_url, found_card_name = None, None, None
#     logging.info(f"シートの列一覧：{df.columns.tolist()}")
#     for col in df.columns:
#         col_str = str(col).strip().lower().replace(" ", "").replace("_", "")
#         target_tweet_id_clean = target_tweet_id.lower().replace(" ", "").replace("_", "")
#         target_url_clean = target_url.lower().replace(" ", "").replace("_", "")
#         target_card_name_clean = target_card_name.lower().replace(" ", "").replace("_", "")
#         if target_tweet_id_clean in col_str:
#             found_tweet_id = col
#         if target_url_clean in col_str:
#             found_url = col
#         if target_card_name_clean in col_str:
#             found_card_name = col
#     if found_tweet_id:
#         logging.info(f"「ツイートID」列を検出:'{found_tweet_id}'")
#     else:
#         logging.warning(f"'{target_tweet_id}'列が見つかりません。")
#     if found_url:
#         logging.info(f"ウェブサイトのURLを検出: '{found_url}'")
#     else:
#         logging.warning(f"'{target_url}'列が見つかりません。")
#     if found_card_name:
#         logging.info(f"カード名を検出: '{found_card_name}'")
#     else:
#         logging.warning(f"'カード名'列が見つかりません。")
#     return found_tweet_id, found_url, found_card_name

def tw_find_columns(df, target_tweet_id=TW_TWEET_ID_COL, target_url=TW_URL_COL, target_card_name="カード名"):
    """Chỉ tìm đúng tên cột, không tìm gần giống."""
    found_tweet_id, found_url, found_card_name = None, None, None
    logging.info(f"シートの列一覧：{df.columns.tolist()}")
    for col in df.columns:
        if col == target_tweet_id:
            found_tweet_id = col
        if col == target_url:
            found_url = col
        if col == target_card_name:
            found_card_name = col
    if found_tweet_id:
        logging.info(f"「ツイートID」列を検出:'{found_tweet_id}'")
    else:
        logging.warning(f"'{target_tweet_id}'列が見つかりません。")
    if found_url:
        logging.info(f"ウェブサイトのURLを検出: '{found_url}'")
    else:
        logging.warning(f"'{target_url}'列が見つかりません。")
    if found_card_name:
        logging.info(f"カード名を検出: '{found_card_name}'")
    else:
        logging.warning(f"'カード名'列が見つかりません。")
    return found_tweet_id, found_url, found_card_name


def tw_process_excel_files(directory, pattern=TW_SOURCE_PATTERN, extension=TW_SOURCE_EXTENSION):
    """Process Excel files containing the pattern in their names for Twitter."""
    search_pattern = os.path.join(directory, f"*{pattern}*{extension}")
    excel_files = [f for f in glob.glob(search_pattern) if not os.path.basename(f).startswith("~$")]

    if not excel_files:
        logging.warning(f"{directory} 内に、パターン「{pattern}」、拡張子「{extension}」のファイルが見つかりません。")
        return []

    all_data = []
    for file_path in excel_files:
        logging.info(f"Excelファイルを処理中：{file_path}")
        try:
            df = pd.read_excel(file_path, sheet_name=TW_SHEET_NAME, engine='openpyxl', header=TW_HEADER_ROW)
            logging.info(f"シート「{TW_SHEET_NAME}」の先頭5行のデータ:\n{df.head(5).to_string()}")
            
            tweet_id_col, url_col, card_name_col = tw_find_columns(df, TW_TWEET_ID_COL, TW_URL_COL)

            if tweet_id_col is None:
                logging.error(f"{os.path.basename(file_path)} 内に「ツイートID」列が見つかりません。")
                continue

            required_cols = [tweet_id_col]
            if url_col:
                required_cols.append(url_col)
            if card_name_col:
                required_cols.append(card_name_col)
            
            df_filtered = df[required_cols]

            tweet_url_card_pairs = set()
            for _, row in df_filtered.iterrows():
                if tweet_id_col in row and not pd.isna(row[tweet_id_col]) and str(row[tweet_id_col]).strip().startswith('i'):
                    tweet_id = str(row[tweet_id_col]).strip()
                    url = str(row[url_col]).strip() if url_col and url_col in df.columns and not pd.isna(row[url_col]) else ""
                    card_name = str(row[card_name_col]).strip() if card_name_col and card_name_col in df.columns and not pd.isna(row[card_name_col]) else ""
                    tweet_url_card_pairs.add((tweet_id, url, card_name))

            unique_pairs = list(tweet_url_card_pairs)
            logging.info(f"{os.path.basename(file_path)} 内で、有効な「ツイートID-URL-カード名」のペアを {len(unique_pairs)} 件検出しました。")
            all_data.extend(unique_pairs)

        except Exception as e:
            logging.error(f"{file_path} の処理中にエラー：{e}")
            continue

    return all_data

def tw_process_csv_files(directory, pattern=TW_CSV_PATTERN, extension=TW_CSV_EXTENSION):
    """Process CSV files containing the pattern in their names for mapping for Twitter."""
    search_pattern = os.path.join(directory, f"*{pattern}*{extension}")
    csv_files = [f for f in glob.glob(search_pattern) if not os.path.basename(f).startswith("~$")]
    
    if not csv_files:
        logging.warning(f"Can't find file  '{pattern}' with '{extension}' in {directory}")
        return []
    
    mapping_data = []
    for file_path in csv_files:
        logging.info(f"紐付け用CSVファイルを処理中：{file_path}")
        try:
            df = pd.read_csv(file_path, encoding='utf-8-sig', usecols=TW_REQUIRED_CSV_COLS)
            
            missing_cols = [col for col in TW_REQUIRED_CSV_COLS if col not in df.columns]
            if missing_cols:
                logging.warning(f"{os.path.basename(file_path)} に列が不足しています：{missing_cols}")
                continue
            
            for _, row in df.iterrows():
                tweet_id = str(row["ツイートID"]).strip() if not pd.isna(row["ツイートID"]) else None
                if tweet_id and tweet_id.startswith('CA-'):
                    mapping_data.append({
                        "tweet_id": tweet_id,
                        "アカウントID": str(row["アカウントID"]).strip() if not pd.isna(row["アカウントID"]) else "",
                        "キャンペーンID": str(row["キャンペーンID"]).strip() if not pd.isna(row["キャンペーンID"]) else "",
                        "キャンペーン名": str(row["広告キャンペーン名"]).strip() if not pd.isna(row["広告キャンペーン名"]) else "",
                        "広告グループID": str(row["広告グループID"]).strip() if not pd.isna(row["広告グループID"]) else "",
                        "広告グループ名": str(row["広告グループ名"]).strip() if not pd.isna(row["広告グループ名"]) else "",
                        "relation ID": str(row["relationID"]).strip() if not pd.isna(row["relationID"]) else ""
                    })
        except Exception as e:
            logging.error(f"{file_path} の処理中にエラー：{e}")
            continue
    
    return mapping_data

def tw_save_unmapped_data(unmapped_data, output_dir, current_date, filename_prefix="missingTWID"):
    """Save unmapped data to yymmdd_missingTWID_Template2.csv or similar."""
    if not os.access(output_dir, os.W_OK):
        logging.error(f"ディレクトリへの書き込み権限がありません：{output_dir}")
        return

    output_file = os.path.join(output_dir, f"{current_date}_{filename_prefix}_Template2.csv")
    if unmapped_data:
        try:
            unmapped_df = pd.DataFrame(unmapped_data, columns=TW_OUTPUT_COLUMNS)
            unmapped_df.to_csv(output_file, index=False, encoding='utf-8-sig')
            logging.info(f"{filename_prefix} のデータ（{len(unmapped_data)} 行）をファイルに保存しました：{output_file}")
        except Exception as e:
            logging.error(f"ファイル書き込みエラー：{output_file} - {e}")
    else:
        logging.info(f"{filename_prefix} に保存するデータがありません。")

def tw_process_files(directory, master_path, output_dir, cid_map):
    """Process Excel, CSV, and Master files to generate output DataFrame for Twitter."""
    current_date = datetime.now().strftime("%y%m%d")
    tw_setup_directories(directory, master_path, output_dir)
    
    mapping_data = tw_process_csv_files(directory)
    excel_data = tw_process_excel_files(directory)
    
    complete_output_data = []
    missing_twid_data = []

    for tweet_id, url, card_name in excel_data:
        modified_id = f"CA-{tweet_id[1:]}"
        matched_mappings = [m for m in mapping_data if m["tweet_id"] == modified_id]
        
        is_incomplete = not matched_mappings or not url or not card_name

        base_output_row = {col: "" for col in TW_OUTPUT_COLUMNS}
        base_output_row.update({
            "Action": TW_ACTION,
            "媒体ID": TW_MEDIA_ID,
            "広告ID": modified_id,
            "パラメ発行済みURL": url,
            "告名": card_name
        })

        if is_incomplete:
            missing_twid_data.append(base_output_row)
            if not matched_mappings:
                logging.warning(f"広告ID {modified_id} の紐付けが見つかりません")
            else:
                logging.warning(f"広告ID {modified_id} は、URLまたは広告名が欠損しているため除外します。")
        else:
            for mapping in matched_mappings:
                account_id = mapping["アカウントID"]
                cid = cid_map.get(account_id, "")
                
                row_to_add = base_output_row.copy()
                row_to_add.update({
                    "CID": cid,
                    "アカウントID": account_id,
                    "キャンペーンID": mapping["キャンペーンID"],
                    "キャンペーン名": mapping["キャンペーン名"],
                    "広告グループID": mapping["広告グループID"],
                    "広告グループ名": mapping["広告グループ名"],
                    "relation ID": mapping["relation ID"]
                })
                
                complete_output_data.append(row_to_add)
                
                if cid:
                    logging.info(f"広告ID {modified_id} をCID {cid} に紐付けました。")
                else:
                    logging.warning(f"広告ID {modified_id} は紐付け済みですが、CIDがありません（アカウントID: {account_id}）。")

    tw_save_unmapped_data(missing_twid_data, output_dir, current_date, "missingTWID")
    
    output_df = pd.DataFrame(complete_output_data, columns=TW_OUTPUT_COLUMNS)
    logging.info(f"データが不完全だった {len(missing_twid_data)} 行を最終出力から除外しました。")
    
    return output_df

# --- CÁC HÀM XỬ LÝ CHO LINE ---

def line_setup_directories():
    """Check and create output directory if needed for Line."""
    for directory in [LINE_INPUT_DIR, os.path.dirname(LINE_MASTER_FILE)]:
        if not os.path.exists(directory):
            logging.error(f"フォルダが見つかりません。: {directory}")
            raise FileNotFoundError(f"フォルダが見つかりません。: {directory}")
    os.makedirs(LINE_OUTPUT_DIR, exist_ok=True)
    if not os.access(LINE_OUTPUT_DIR, os.W_OK):
        logging.error(f"Không có quyền ghi vào thư mục: {LINE_OUTPUT_DIR}")
        raise PermissionError(f"Không có quyền ghi vào thư mục: {LINE_OUTPUT_DIR}")

def line_read_excel_file(file_path, sheet_name=None, header_row=0, required_cols=None):
    """Read an Excel file, check its existence, and load only required columns for Line."""
    if not os.path.exists(file_path):
        logging.error(f"File '{file_path}' not found. Please check the file name or path.")
        return None, None
    try:
        workbook = openpyxl.load_workbook(file_path)
        logging.info(f"Các sheet in file '{os.path.basename(file_path)}': {workbook.sheetnames}")
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, usecols=required_cols)
        logging.info(f"Tổng số line in sheet '{sheet_name}' (bao gồm tiêu đề): {len(df) + 1}")
        logging.info(f"Các cột in sheet '{sheet_name}': {df.columns.tolist()}")
        return df, workbook.sheetnames
    except Exception as e:
        logging.error(f"error read file '{file_path}': {e}")
        return None, None

def line_read_csv_file(file_path, required_cols=None):
    """Read a CSV file, trying multiple encodings and handling errors for Line."""
    if not os.path.exists(file_path):
        logging.error(f"File '{file_path}' not found. Please check the file name or path.")
        return None
    encodings = ['utf-8-sig', 'utf-16', 'shift-jis', 'latin1']
    for encoding in encodings:
        try:
            df = pd.read_csv(file_path, encoding=encoding, sep='\t', on_bad_lines='warn', dtype=str, usecols=required_cols)
            logging.info(f"Tổng số line in file '{os.path.basename(file_path)}' (bao gồm tiêu đề): {len(df) + 1}")
            logging.info(f"Các cột in file '{os.path.basename(file_path)}': {df.columns.tolist()}")
            logging.info(f"read file '{os.path.basename(file_path)}' sucssecful with encoding: {encoding}")
            return df
        except UnicodeDecodeError as e:
            logging.warning(f"can't read file '{os.path.basename(file_path)}' với encoding {encoding}: {e}")
            continue
        except pd.errors.ParserError as e:
            logging.warning(f"Lỗi phân tích file '{os.path.basename(file_path)}' với encoding {encoding} và sep='\\t': {e}")
            try:
                df = pd.read_csv(file_path, encoding=encoding, sep='\t', on_bad_lines='skip', dtype=str, usecols=required_cols)
                logging.info(f"Tổng số line in file '{os.path.basename(file_path)}' (bao gồm tiêu đề): {len(df) + 1}")
                logging.info(f"Các cột in file '{os.path.basename(file_path)}': {df.columns.tolist()}")
                logging.info(f"read file '{os.path.basename(file_path)}' sucssecful with encoding: {encoding} (bỏ qua line lỗi)")
                return df
            except Exception as e2:
                logging.warning(f"Still can't read file với encoding {encoding} và sep='\\t': {e2}")
                continue
    logging.error(f"can't read file '{file_path}' với bất kỳ encoding nào in {encodings}.")
    return None

def line_check_columns(df, required_cols, source_name):
    """Check if required columns exist in the DataFrame for Line."""
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        logging.error(f"No rule in {source_name}: {missing_cols}")
        raise ValueError(f"Missing columns in {source_name}: {missing_cols}")

def line_check_duplicates(df, column, file_path, log_file):
    """Check for duplicates in a column and log them for Line."""
    duplicates = df[df[column].duplicated(keep=False)]
    if not duplicates.empty:
        logging.warning(f"Có {len(duplicates)} Duplicate values in rule {column}")
        logging.warning(f"Duplicate values: {duplicates[column].unique().tolist()}")
        duplicates.to_csv(log_file, index=False, encoding='utf-8-sig')
        logging.info(f"Duplicate values save in: {log_file}")

def line_process_files():
    """Process files for Line and return output DataFrame."""
    line_setup_directories()
    
    logging.info("\nRead Master.xlsx:")
    df_master, _ = line_read_excel_file(LINE_MASTER_FILE, sheet_name='Sheet1', header_row=0, required_cols=LINE_MASTER_REQUIRED_COLS)
    if df_master is None:
        raise SystemExit("Can not continue with read Master.xlsx")
    line_check_columns(df_master, LINE_MASTER_REQUIRED_COLS, "sheet 'Sheet1' của Master.xlsx")

    logging.info("\nFind file Master2:")
    master2_files = [f for f in os.listdir(LINE_INPUT_DIR) if 'campaign-adgroup-ad' in f.lower() and f.endswith('.csv')]
    if not master2_files:
        raise SystemExit(f"Can't find file .csv  'campaign-adgroup-ad' in  {LINE_INPUT_DIR}")
    
    df_master2_list = []
    for file in master2_files:
        master2_file = os.path.join(LINE_INPUT_DIR, file)
        df = line_read_csv_file(master2_file)
        if df is not None:
            df_master2_list.append(df)
    
    if not df_master2_list:
        raise SystemExit("No 掲載内容.")
    
    df_master2 = pd.concat(df_master2_list, ignore_index=True)
    line_check_columns(df_master2, LINE_MASTER2_REQUIRED_COLS, "Master2")
    df_master2['combined_key'] = (df_master2['Campaign name'].astype(str) + '|' + 
                                  df_master2['Ad group name'].astype(str) + '|' + 
                                  df_master2['Ad name'].astype(str))
    if 'Start date' in df_master2.columns:
        df_master2 = df_master2.sort_values('Start date', ascending=False).drop_duplicates('combined_key', keep='first')

    logging.info("\nFind file input:")
    input_files = [f for f in os.listdir(LINE_INPUT_DIR) if '入稿' in f and f.endswith('.xlsm')]
    if not input_files:
        raise SystemExit(f"Can't find file .xlsm  '入稿' in this foder {LINE_INPUT_DIR}")

    df_input_list = []
    for file in input_files:
        input_file = os.path.join(LINE_INPUT_DIR, file)
        df, _ = line_read_excel_file(input_file, sheet_name='AD入稿', header_row=14, required_cols=LINE_INPUT_REQUIRED_COLS)
        if df is not None:
            df_input_list.append(df.dropna(subset=LINE_INPUT_REQUIRED_COLS, how='all'))

    if not df_input_list:
        raise SystemExit("no file.")

    df_input = pd.concat(df_input_list, ignore_index=True)
    df_input['combined_key'] = (df_input['キャンペーン名'].astype(str) + '|' + 
                                df_input['広告グループ名'].astype(str) + '|' + 
                                df_input['広告名'].astype(str))

    df_output = pd.DataFrame()
    df_output['アカウントID'] = df_input['アカウントID']
    df_output['キャンペーン名'] = df_input['キャンペーン名']
    df_output['広告グループ名'] = df_input['広告グループ名']
    df_output['告名'] = df_input['広告名']
    df_output['パラメ発行済みURL'] = df_input['リンク先URL1']
    df_output['combined_key'] = df_input['combined_key']

    df_output = df_output.merge(df_master[['アカウントID', 'CID']].drop_duplicates(), on='アカウントID', how='left')
    df_output = df_output.merge(
        df_master2[['combined_key', 'Campaign ID', 'Ad group ID', 'Ad ID']].drop_duplicates(), 
        on='combined_key', 
        how='left'
    )
    df_output.drop(columns=['combined_key'], inplace=True, errors='ignore')
    df_output.rename(columns={'Campaign ID': 'キャンペーンID', 'Ad group ID': '広告グループID', 'Ad ID': '広告ID'}, inplace=True)

    for col in LINE_OUTPUT_COLUMNS:
        if col not in df_output.columns:
            df_output[col] = ''
    
    for col in df_output.select_dtypes(include=['object']).columns:
        df_output[col] = df_output[col].fillna('')

    df_output['Action'] = LINE_ACTION
    df_output['媒体ID'] = LINE_MEDIA_ID

    missing_mask = (df_output['キャンペーンID'] == '') | \
                   (df_output['広告グループID'] == '') | \
                   (df_output['広告ID'] == '')
    
    missing_ids_df = df_output[missing_mask]

    if not missing_ids_df.empty:
        logging.warning(f"Có {len(missing_ids_df)} line in output thiếu ít nhất một in các ID (キャンペーンID, 広告グループID, 広告ID)")
        missing_ids_df[['キャンペーン名', '広告グループ名', '告名']].to_csv(LINE_MISSING_IDS_FILE, index=False, encoding='utf-8-sig')
        logging.info(f"line thiếu ID được lưu in: {LINE_MISSING_IDS_FILE}")

    df_output_final = df_output[~missing_mask]
    logging.info(f"IDが欠損していた {len(missing_ids_df)} 行を最終出力から除外しました。")
    
    df_output_final = df_output_final[LINE_OUTPUT_COLUMNS]
    logging.info(f"Số line in file output: {len(df_output_final)}")
    
    return df_output_final

# CÁC HÀM XỬ LÝ CHO FACEBOOK

def process_facebook_data(input_dir, output_dir, cid_map):
    search_path = os.path.join(input_dir, f"*{FB_CSV_PATTERN}*{FB_CSV_EXTENSION}")
    fb_files = glob.glob(search_path)
    
    if not fb_files:
        logging.info("FacebookのエクスポートCSVファイルが見つかりません。")
        return None

    # ✅ Khởi tạo list tổng để chứa kết quả từ TẤT CẢ các file
    all_processed_rows = []
    all_app_store_link_rows = []

    # 🔄 Bắt đầu vòng lặp để xử lý từng file tìm được
    for file_path in fb_files:
        logging.info(f"Facebookファイルを処理中：{file_path}")

        df = None
        # Thử đọc file với các encoding khác nhau
        encodings_to_try = ['utf-8-sig', 'utf-16', 'shift_jis', 'cp932', 'utf-8']
        for encoding in encodings_to_try:
            try:
                df = pd.read_csv(file_path, sep='\t', engine='python', on_bad_lines='warn', dtype=str, encoding=encoding)
                logging.info(f"ファイル「{os.path.basename(file_path)}」の読み込みに成功しました（エンコーディング：{encoding}）。")
                break
            except Exception:
                continue

        if df is None:
            logging.error(f"ファイル「{os.path.basename(file_path)}」を読み込めません。スキップします。")
            continue # Chuyển sang file tiếp theo

        try:
            # Định nghĩa các chuỗi link app store cần lọc
            google_play_url = "http://play.google.com/store/apps"
            apple_store_url = "http://itunes.apple.com/app"
            
            extra_link_cols = [f'Product {i} - Link' for i in range(1, 10)] + \
                              [f'Additional Link {i}' for i in range(1, 10)]
            existing_extra_link_cols = [col for col in extra_link_cols if col in df.columns]

            # Lặp qua từng dòng trong file CSV hiện tại
            for index, row in df.iterrows():
                row_account_id = ''
                row_cid = 'empty'
                current_hash = row.get('Image Hash') if pd.notna(row.get('Image Hash')) else ''

                if current_hash and ':' in current_hash:
                    row_account_id = f"CA-{current_hash.split(':')[0]}"
                    row_cid = cid_map.get(row_account_id, "empty")
                
                main_link = row.get('Link', '')

                base_data = {
                    "Action": FB_ACTION, 
                    "媒体ID": FB_MEDIA_ID, 
                    "アカウントID": row_account_id,
                    "CID": row_cid,
                    'キャンペーンID': str(row.get('Campaign ID', '')).replace('cg:', 'CA-'),
                    'キャンペーン名': str(row.get('Campaign Name', '')),
                    '広告グループID': str(row.get('Ad Set ID', '')).replace('c:', 'CA-'),
                    '広告グループ名': str(row.get('Ad Set Name', '')),
                    '広告ID': str(row.get('Ad ID', '')).replace('a:', 'CA-'),
                    '告名': str(row.get('Ad Name', '')),
                    'relation ID': '', 'キーワードID': '', 'キーワード名': '', 'マッチタイプ': '', 'ドラフト停止日': ''
                }
                
                # 1. Xử lý Main link
                if pd.notna(main_link) and main_link.strip() != '':
                    main_row = base_data.copy()
                    main_row['パラメ発行済みURL'] = main_link
                    if google_play_url in main_link or apple_store_url in main_link:
                        all_app_store_link_rows.append(main_row)
                    else:
                        all_processed_rows.append(main_row)

                # 2. Xử lý các link phụ
                processed_extra_links = set() 
                for col_name in existing_extra_link_cols:
                    extra_link = row.get(col_name, '')
                    if pd.notna(extra_link) and extra_link.strip() != '' and extra_link != main_link and extra_link not in processed_extra_links:
                        processed_extra_links.add(extra_link)
                        extra_row = base_data.copy()
                        extra_row['パラメ発行済みURL'] = extra_link
                        if google_play_url in extra_link or apple_store_url in extra_link:
                            all_app_store_link_rows.append(extra_row)
                        else:
                            all_processed_rows.append(extra_row)
        
        except Exception as e:
            logging.error(f"ファイル「{os.path.basename(file_path)}」のデータ処理中にエラー：{e}")
            continue # Bỏ qua file lỗi và tiếp tục với file khác   

    # --- KẾT THÚC VÒNG LẶP ---

    # ✅ Bây giờ, xử lý kết quả tổng hợp từ tất cả các file
    if all_app_store_link_rows:
        missing_df = pd.DataFrame(all_app_store_link_rows, columns=FB_OUTPUT_COLUMNS)
        missing_file_path = os.path.join(output_dir, 'missingFacebookID_Template2.csv')
        missing_df.to_csv(missing_file_path, index=False, encoding='utf-8-sig')
        logging.info(f"アプリストアへのリンクを含む {len(missing_df)} 行をファイルに保存しました：{missing_file_path}")

    if not all_processed_rows:
        logging.warning("全てのFacebookファイルに、処理可能な有効な行（アプリリンク以外）はありませんでした。")
        return pd.DataFrame(columns=FB_OUTPUT_COLUMNS)

    fb_results_df = pd.DataFrame(all_processed_rows, columns=FB_OUTPUT_COLUMNS)
    logging.info(f"処理完了。{len(fb_files)}個のファイルから合計{len(fb_results_df)}件の有効なデータを取得しました。")
    return fb_results_df

# --- CÁC HÀM CHUNG ---

def get_and_validate_username():
    try:
        df = pd.read_excel(user_file)
        if df.empty:
            logging.error(f"エラー：Usernameをみつかりません。")
            sys.exit() 

        username = str(df.iloc[0, 0])

        # ĐIỀU KIỆN KIỂM TRA MỚI
        if username == "DT06": 
            logging.error(f"エラー：人間アクセス。。 ")
            sys.exit()

        logging.info(f"Username '{username}' Verified ✅")
        return username

    except FileNotFoundError:
        logging.error(f"エラー：Usernameをみつかりません。 {user_file}")
        sys.exit() # Dừng chương trình
    except Exception as e:
        logging.error(f"エラー：Usernameをみつかりません。 {e}")
        sys.exit() # Dừng chương trình



def check_encoding(file_path):
    """Kiểm tra encoding của file."""
    try:
        with open(file_path, 'rb') as file:
            result = detect(file.read())
            encoding = result['encoding']
            return encoding
    except Exception as e:
        logging.error(f"Error when check encoding of {file_path}: {e}")
        return None

def read_csv_with_fallback(file_path, expected_encoding=None, sep=','):
    """read file CSV với các encoding thử nghiệm và dấu phân tách."""
    encodings = [expected_encoding] if expected_encoding else ['utf-8-sig', 'utf-8', 'utf-16', 'latin1', 'cp1252']
    for encoding in encodings:
        try:
            df = pd.read_csv(file_path, encoding=encoding, sep=sep)
            logging.info(f"read file {file_path} sucssecful with encoding {encoding} and '{sep}'")
            return df, encoding
        except UnicodeDecodeError:
            logging.info(f"can't read file {file_path} with {encoding}, try next encoding...")
        except Exception as e:
            logging.error(f"Eror when read file {file_path} với {encoding}: {e}")
    raise ValueError(f"can't read file {file_path} with any encoding in {encodings}")

def merge_with_template2(input_df, template_file_path, output_dir, media):
    """入力DataFrameをTemplate.xlsxとマージし、CIDごとにExcelファイルとして分割して出力する。"""
    try:
        df = pd.read_excel(user_file)
        username = get_and_validate_username()

        # 1. Excelテンプレートファイルの存在確認
        if not os.path.exists(template_file_path):
            raise FileNotFoundError(f"テンプレートファイルが見つかりません: {template_file_path}")

        # 2. Excelファイルを読み込む
        logging.info(f"テンプレートファイルを読み込んでいます: {template_file_path}")
        template_df = pd.read_excel(template_file_path, engine='openpyxl', dtype=str)
        template_df = template_df.fillna('')

        logging.info("テンプレートファイルの列:")
        logging.info(template_df.columns.tolist())
        
        # 3. 入力データとテンプレートの列を比較
        if not all(col in template_df.columns for col in input_df.columns):
            missing_cols = [col for col in input_df.columns if col not in template_df.columns]
            raise ValueError(f"入力データの列がテンプレートと一致しません！不足している列: {missing_cols}")
        
        # 4. データフレームを結合
        combined_df = pd.concat([template_df, input_df], ignore_index=True)
        
        current_date = datetime.now().strftime("%y%m%d")
        combined_df['CID'] = combined_df['CID'].astype(str).replace('nan', '').replace('.0', '')
        
        # 5. CIDごとにグループ化してファイルを出力
        cid_groups = combined_df.groupby('CID')
        
        logging.info("CIDに基づいてファイルを分割・Excel形式で出力します...")
        for cid, group in cid_groups:
            cid_prefix = str(cid).strip() if cid and len(str(cid).strip()) >= 6 else 'empty'
            
            # ▼ 変更点 1: 出力ファイル名の拡張子を .xlsx に変更
            output_file = os.path.join(output_dir, f'{username}★{current_date}_{cid_prefix}_{media}_Template2.xlsx')
            
            try:
                # ▼ 変更点 2: to_excel() を使ってExcelファイルとして保存
                group.to_excel(output_file, index=False, engine='openpyxl')
                logging.info(f"CID '{cid}' の出力ファイル ({len(group)}行) を保存しました: {output_file}")
            except PermissionError:
                logging.error(f"書き込み権限がありません: {output_file}")
        
    except FileNotFoundError as e:
        logging.error(f"エラー: {e}")
    except ValueError as e:
        logging.error(f"エラー: {e}")
    except Exception as e:
        logging.error(f"不明なエラーが発生しました: {e}", exc_info=True)
        

def run_filerun():
    try:
        # 1. Lấy thông tin ngày giờ hiện tại
        now = datetime.now()
        yyyymmdd = now.strftime('%Y%m%d')
        month_folder_name = now.strftime('%m月')

        try:
            df = pd.read_excel(user_file)
            username = get_and_validate_username()
            
            if not username:
                print("エラー: Usernameをみつかりません。")
                return 
            
            print(f"Username情報を読み済です。. Username: {username}")

        except FileNotFoundError:
            print(f"エラー: '{user_file}'をみつかりません。.STOP")
            return # Hợp lệ
        except Exception as e:
            print(f"Excelファイルの読み取り中に不明なエラーが発生しました {e}.")
            return # Hợp lệ

        # 3. Tạo đường dẫn đích hoàn chỉnh  \\vnfs\share\04_SEM\007.PARAMETER_STORAGE   C:\Users\s14335\Desktop\truong\04_SEM
        destination_base = r'\\vnfs\share\04_SEM\007.PARAMETER_STORAGE' 
        destination_folder = os.path.join(
            destination_base, 
            month_folder_name, 
            yyyymmdd, 
            f"{yyyymmdd}★{username}"
        )

        # In ra các đường dẫn để kiểm tra
        print(f"Source directory: {source_folder}")
        print(f"Destination folder: {destination_folder}")

        # Kiểm tra và tạo thư mục đích nếu nó chưa tồn tại
        if not os.path.exists(destination_folder):
            os.makedirs(destination_folder)
            print(f"Destination folder created {destination_folder}")

        files_in_source = os.listdir(source_folder)

        if not files_in_source:
            print(f"Source directory '{source_folder}' empty, can't copy file")
            return

        print(f"\nCopy starting...")
        copied_count = 0
        
        for file_name in files_in_source:
            source_path = os.path.join(source_folder, file_name)
            if os.path.isfile(source_path):
                if "empty" not in file_name and "missing" not in file_name:
                    destination_path = os.path.join(destination_folder, file_name)
                    shutil.copy2(source_path, destination_path)
                    print(f" ✅ Copied: {file_name}")
                    copied_count += 1
                else:
                    print(f" ❌ Ignore: {file_name} (contains 'empty' or 'missing')")

        print(f"\n Done! Copied {copied_count} file. ✅")
    
    except Exception as e:
        print(f"An unexpected error occurred while copying the file: {e}")    
        
        
def upload_data_to_google_sheet(processed_df):
    """
    Trích xuất thông tin từ DataFrame đã xử lý và dán lên Google Sheet.
    """
    if processed_df is None or processed_df.empty:
        logging.warning("Không có dữ liệu để tải lên Google Sheet.")
        return

    print("\n=====================================================================")
    print("=== ステップ➌：Google Sheetsへのデータアップロード開始 ===")
    print("=====================================================================")
    
    try:
        # 1. Xác thực
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
        creds_path = os.path.join(base_dir, 'credentials.json')
        creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
        client = gspread.authorize(creds)

        # 2. Mở file Google Sheet và chọn sheet "BOT"
        spreadsheet_url = "https://docs.google.com/spreadsheets/d/1sw_eGCeogGVz9q_yqiDDhUVdWXM1AHgrhObQdTgyDU0/edit?gid=0#gid=0"
        spreadsheet = client.open_by_url(spreadsheet_url)
        worksheet = spreadsheet.worksheet("BOT")
        print(f"✅ Đã kết nối thành công tới Google Sheet, sheet 'BOT'.")

        # 3. Chuẩn bị dữ liệu để tải lên
        creation_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        username = get_and_validate_username()
        
        # Lấy danh sách account_id duy nhất từ DataFrame
        if 'アカウントID' not in processed_df.columns:
            logging.error("Không tìm thấy cột 'アカウントID' trong dữ liệu đã xử lý.")
            return
            
        account_ids = processed_df['アカウントID'].dropna().unique().tolist()
        
        rows_to_append = []
        for acc_id in account_ids:
            if acc_id: # Chỉ thêm nếu account_id không rỗng
                rows_to_append.append([creation_date, username, str(acc_id)])

        if not rows_to_append:
            logging.warning("Không có Account ID hợp lệ để tải lên.")
            return
            
        # 4. Ghi dữ liệu vào cuối sheet
        worksheet.append_rows(rows_to_append, value_input_option='USER_ENTERED')
        
        print(f"✅ Đã tải thành công {len(rows_to_append)} dòng dữ liệu lên Google Sheet.")
        print("====== ステップ➌ 完了 - アップロード成功 ======")

    except FileNotFoundError:
        logging.error("LỖI: Không tìm thấy file 'credentials.json'. Vui lòng kiểm tra lại Phần 1.")
    except gspread.exceptions.SpreadsheetNotFound:
        logging.error("LỖI: Không tìm thấy Google Sheet. Vui lòng kiểm tra lại URL và quyền chia sẻ.")
    except gspread.exceptions.WorksheetNotFound:
        logging.error("LỖI: Không tìm thấy sheet có tên 'BOT' trong file Google Sheet.")
    except Exception as e:
        logging.error(f"Đã xảy ra lỗi không mong muốn khi tải dữ liệu lên Google Sheet: {e}")



def run_parameter_generation():
    """Hàm chính for chạy logic tạo file tham số."""
    options = {1: 'Twitter', 2: 'Line', 3: 'Facebook', 0: '終了'}
    init()
    print(Fore.BLUE + Back.BLACK + Style.NORMAL + r"  _____                                    _____ _                              "+ Style.RESET_ALL)
    print(Fore.BLUE + Back.BLACK + Style.NORMAL + r" |  __ \                                  / ____| |                             "+ Style.RESET_ALL)
    print(Fore.BLUE + Back.BLACK + Style.NORMAL + r" | |__) |_ _ _ __ __ _ _ __ ___   ___    | (___ | |_ ___  _ __ __ _  __ _  ___  "+ Style.RESET_ALL)
    print(Fore.BLUE + Back.BLACK + Style.NORMAL + r" |  ___/ _` | '__/ _` | '_ ` _ \ / _ \    \___ \| __/ _ \| '__/ _` |/ _` |/ _ \ "+ Style.RESET_ALL)
    print(Fore.BLUE + Back.BLACK + Style.NORMAL + r" | |  | (_| | | | (_| | | | | | |  __/    ____) | || (_) | | | (_| | (_| |  __/ "+ Style.RESET_ALL)
    print(Fore.BLUE + Back.BLACK + Style.NORMAL + r" |_|   \__,_|_|  \__,_|_| |_| |_|\___|   |_____/ \__\___/|_|  \__,_|\__, |\___| "+ Style.RESET_ALL)
    print(Fore.BLUE + Back.BLACK + Style.NORMAL + r"                                                                     __/ |      "+ Style.RESET_ALL)
    print(Fore.BLUE + Back.BLACK + Style.NORMAL + r"                                                                    |___/       "+ Style.RESET_ALL)
    print("\n1：Twitter")
    print("2：Line")
    print("3：Facebook")
    print("0：終了")
    choice = -1
    while choice not in options:
        try:
            choice = int(input("媒体を番号で選択ください："))
            if choice not in options:
                logging.warning(f"不正な値: {choice}. ( 1, 2, 3, または 0 を選択してください。)")
        except ValueError:
            logging.warning("整数 (1, 2, 3, または 0) を入力してください。")
            choice = -1 # Reset for vòng lặp tiếp tục

    # Nếu người dùng chọn kết thúc, trả về tín hiệu for dừng hoàn toàn
    if choice == 0:
        logging.info(f"プログラムを終了します。 {options[0]}")
        return (False, False) 

    # --- HỎI VỀ VIỆC UPLOAD NGAY in ĐÂY ---
    upload_choice = ''
    while upload_choice not in ['1', '2']:
        upload_choice = input("\nCLOUDへ自動的にアップロードしますか？ (1: はい, 2: いいえ): ")
        if upload_choice not in ['1', '2']:
            print("不正な選択です。1とか2を入力してください。")
    
    should_upload = (upload_choice == '1')

    # --- CÁC THƯ MỤC CHUNG ---
    source_dir = os.path.join(base_dir, 'input')
    master_dir = os.path.join(base_dir, 'master')
    output_dir = os.path.join(base_dir, 'output')
    
    # --- TIẾP TỤC TẠO FILE ---
    output_df = None
    media = options.get(choice, 'Unknown')
    cid_map = tw_load_master_file(master_dir)

    if choice == 1:
        logging.info(f"{options[1]}の処理を{datetime.now()}に開始しました。")
        output_df = tw_process_files(source_dir, master_dir, output_dir, cid_map)
        if output_df is not None:
             logging.info(f"「{options[1]}」の処理完了。処理データ：{len(output_df)}行")
    elif choice == 2:
        logging.info(f"{options[2]}の処理を{datetime.now()}に開始しました。")
        output_df = line_process_files()
        if output_df is not None:
            logging.info(f"「{options[2]}」の処理完了。処理データ：{len(output_df)}行")
    elif choice == 3:
        logging.info(f"{options[3]}の処理を{datetime.now()}に開始しました。")
        output_df = process_facebook_data(source_dir, output_dir, cid_map)
        if output_df is not None:
            logging.info(f"「{options[3]}」の処理完了。処理データ：{len(output_df)}行")
        else:
            logging.warning("出力するデータがありませんでした。")
            output_df = pd.DataFrame() 
    if output_df is not None and not output_df.empty:
        logging.info("Work with Template2.csv")
        merge_with_template2(output_df, TEMPLATE2_FILE, LINE_OUTPUT_DIR, media)
        logging.info("Work with Template2 done.")
    else:
        logging.warning("No data output,....")

    # Trả về tín hiệu tiếp tục và có upload hay không
    return (output_df, True, should_upload, True)

# Chạy hàm DATA BASE

def log_successful_upload(file_path_to_log):
    try:
        log_df = pd.read_excel(file_path_to_log, engine='openpyxl', dtype=str)

        # Thêm cột Timestamp vào mỗi line
        log_df['Timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cols = log_df.columns.tolist()
        cols = ['Timestamp'] + [col for col in cols if col != 'Timestamp']
        log_df = log_df[cols]

        # Kiểm tra xem file log đã tồn in chưa
        if os.path.exists(UPLOAD_LOG_FILE):
            # Nếu đã tồn in, ghi nối tiếp vào file mà không cần ghi header
                existing_df = pd.read_excel(UPLOAD_LOG_FILE, engine='openpyxl')
                db_columns = existing_df.columns
                log_df_aligned = log_df.reindex(columns=db_columns)
                combined_df = pd.concat([existing_df, log_df_aligned], ignore_index=True)
                combined_df.to_excel(UPLOAD_LOG_FILE, index=False, engine='openpyxl')
        else:
            # Nếu chưa tồn in, tạo file mới và ghi cả header
            log_df.to_excel(UPLOAD_LOG_FILE, index=False, engine='openpyxl')
        
        print(f"💾 Databaseに全て記載されていました！: {os.path.basename(file_path_to_log)}")

    except Exception as e:
        print(f"❗️ Databaseの記載エラーが発生しました。: {e}")

# =================================================================================
# PHẦN 2: AUTO UP
# =================================================================================

def perform_automated_login(driver, username, password, cookie_path):
    """
    Thực hiện đăng nhập hoàn toàn tự động và chỉ lưu cookie khi đã vào đúng trang đích.
    """
    try:
        print("アカウントで自動ログインを開始します...")
        driver.get(URL_LOGIN)

        # Nhấn nút Login ban đầu
        login_button = WebDriverWait(driver, 60).until(
            #EC.element_to_be_clickable((By.XPATH, "//button[@type='Login' and span[text()='ログイン']]"))
            EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Login']"))
            #(By.XPATH, "//button[@data-slot='button' and normalize-space()='Login']")
            #(By.CSS_SELECTOR, "button[data-slot='button']")
        )
        driver.execute_script("arguments[0].click();", login_button)

        # Điền username và password
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.XPATH, "//input[@type='text' or @type='email']"))
        )
        driver.find_element(By.XPATH, "//input[@type='text' or @type='email']").send_keys(username)
        driver.find_element(By.XPATH, "//input[@type='password']").send_keys(password)

        # Nhấn nút Sign On
        sign_on_button = WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.ID, "signOnButton")))
        driver.execute_script("arguments[0].click();", sign_on_button)
        
        # Chờ và kiểm tra xem URL đã đúng là trang đích chưa
        print("ログイン後にリダイレクトを待機しています。。。。")
        # time.sleep(60) # Chờ thêm 60 giây để xử lý các bước xác thực PING ID
        WebDriverWait(driver, 60).until(EC.url_contains("schedule")) # Chờ URL có chứa "schedule"
        # Đảm bảo trang được tải lại và chuyển đến đúng URL một cách mạnh mẽ hơn
        for _ in range(3):
            try:
                driver.get(URL_TARGET)
                # driver.refresh()
                WebDriverWait(driver, 2).until(EC.url_to_be(URL_TARGET))
                break
            except Exception:
                time.sleep(2)
        else:
            print("ページのリフレッシュまたは遷移に失敗しました。")
        WebDriverWait(driver, 5).until(EC.url_to_be(URL_TARGET)) # Chờ chính xác URL đích

        # Nếu đã đến được đây, nghĩa là đăng nhập 100% thành công
        print("✅ 自動ログインとリダイレクトが成功しました！")
        print("新しい Cookie を保存しています。。。")
        
        # Thêm 1 giây chờ để đảm bảo mọi cookie đã được trình duyệt thiết lập
        time.sleep(1) 
        
        with open(cookie_path, 'w') as file:
            json.dump(driver.get_cookies(), file)
        print(f"💾 新しい Cookie が保存されました: {cookie_path}")
        return True

    except Exception as e:
        print(f"❌ 自動ログイン中にエラーが発生しました: {e}")
        # In ra URL hiện tại để gỡ lỗi
        try:
            print(f"記録された最終 URL: {driver.current_url}")
        except:
            pass
        return False

def run_auto_uploader():
    # === GIAI ĐOẠN 1: KHỞI TẠO VÀ LẤY THÔNG TIN ===
    try:
        df = pd.read_excel(user_file)
        username = get_and_validate_username()
        password = str(df.iloc[0, 1])
        if not username or not password:
            print("エラー:Excelファイルでユーザー名またはパスワードが見つかりません。アップロード処理を終了します。")
            return
        print(f"ログイン情報の読み込みに成功しました。ユーザー名： {username}")

    except Exception as e:
        print(f"エラー：指定されたパスにファイルが見つかりません。 {e}")
        return

    output_directory = os.path.join(base_dir, 'output')
    files_to_upload = [os.path.join(output_directory, f) for f in os.listdir(output_directory) if f.endswith('.xlsx') and '_empty_' not in f and 'missing' not in f]

    if not files_to_upload:
        print("アップロード対象の有効なファイルが見つかりません。アップロード処理を終了します。")
        return

    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    login_successful = False
    
    try:
        # === GIAI ĐOẠN 2: THỰC HIỆN ĐĂNG NHẬP ===
        # Ưu tiên đăng nhập bằng cookie
        if os.path.exists(cookie_path):
            print("\nCookie ファイルが見つかりました。自動的にログインしようとしています...")
            driver.get(URL_BASE)
            with open(cookie_path, "r") as file:
                cookies = json.load(file)
            for cookie in cookies:
                try:
                    driver.add_cookie(cookie)
                except Exception:
                    # Bỏ qua cookie lỗi
                    pass
            
            driver.get(URL_TARGET)
            time.sleep(3)
            
            if "login" not in driver.current_url:
                print("Cookie を使用したログインに成功しました!")
                login_successful = True
            else:
                print("Cookieの有効期限が切れています。アカウントで再度ログインしています。。。")
                os.remove(cookie_path)
        
        # Nếu đăng nhập cookie không thành công, đăng nhập bằng tài khoản
        if not login_successful:
            login_successful = perform_automated_login(driver, username, password, cookie_path)

        # === GIAI ĐOẠN 3: TẢI FILE LÊN (CHỈ CHẠY KHI ĐĂNG NHẬP THÀNH CÔNG) ===
        if login_successful:
            print("\nファイルのアップロードを開始します...")
            
            # --- ĐOẠN MÃ UPLOAD CỦA BẠN ĐƯỢC THÊM VÀO ĐÂY ---
            for index, file_path in enumerate(files_to_upload):
                filename = os.path.basename(file_path)
                print(f"\n--- [{index + 1}/{len(files_to_upload)}] {filename}: 処理を開始---")

                upload_successful = False
                retry_count = 0

                while not upload_successful and retry_count < 3: # Thêm giới hạn 3 lần thử
                    try:
                        # Mỗi lần thử lại, ta quay lại trang Manager
                        driver.refresh()
                        driver.get("https://ca-rpa.cloud/en/parameter-storage/manager")
                        time.sleep(1)
                        # Đợi trang tải xong
                        WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.XPATH, "//input[@role='combobox']"))
                        )
                        time.sleep(1) 

                        # 1. Tìm kiếm với CID
                        search_cid = filename.split('_')[1][:6]
                        print(f"➞[第 {retry_count + 1}回] '{search_cid}'CIDを選択します。。。")
                        WebDriverWait(driver, 10).until(EC.url_contains("parameter-storage/manager"))
                        search_input = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//input[@role='combobox']")))
                        search_input.clear()
                        search_input.send_keys(search_cid)

                        # 2. Nhấp vào kết quả tìm kiếm
                        print("➞CIDを入力")
                        search_result_xpath = f"//*[starts-with(normalize-space(.), '{search_cid}') and contains(@class, 'option')]"
                        search_result_element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, search_result_xpath)))
                        search_result_element.click()
                        time.sleep(1)

                        # 3. Tải file lên
                        print("➞ファイルをアップロード")
                        upload_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[@type='file']")))
                        upload_input.send_keys(file_path)
                        time.sleep(1)

                        # 4. Nhấn nút "Upload Data"
                        upload_buttons = driver.find_elements(By.XPATH, "//button[.//span[text()='Upload']]")
                        if upload_buttons:
                            btn = upload_buttons[-1]  
                            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", btn)
                            time.sleep(1)
                            driver.execute_script("arguments[0].click();", btn)
                            print("➞'Upload Data'を押しました。")
                        
                            confirm_btn = WebDriverWait(driver, 15).until(
                                EC.element_to_be_clickable((
                                By.XPATH,
                                "//button[.//span[normalize-space(text())='Confirm Upload'] or contains(., 'Confirm Upload')]"
                            )) 
                            )
                            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", confirm_btn)
                            time.sleep(1)
                            driver.execute_script("arguments[0].click();", confirm_btn)
                            print("➞'Confirm Upload'を押しました。")
                        else:
                            print("Không tìm thấy nút Upload.")
                            raise Exception("Upload button not found.")
                        
                        # # 5. Nhấn nút "Cancel" trên popup
                        
                        # cancel_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Stay on this page']")))
                        # cancel_button.click()
                        
                        log_successful_upload(file_path)

                        upload_successful = True
                        print(f"🟢 {filename}をアップロード成功しました！")
                        time.sleep(1)  

                    except Exception as file_error:
                        retry_count += 1
                        print(f" {filename}:操作中にエラーが発生しました {file_error}")
                        print(f"   -> 5秒後に再試行します。 ({retry_count}回目の試行 )")
                        time.sleep(2)
            # --- KẾT THÚC ĐOẠN MÃ UPLOAD ---

            print("\n==============================================")
            print("すべてのファイルの処理が完了しました！")
            print("==============================================")
        else:
            print("\nログイン失敗した。アップ不可です。")

    except Exception as e:
        print(f"\nアップロード処理中のエラー: {e}")
    finally:
        time.sleep(2)
        print("ブラウザーを終了します。")
        driver.quit()

# =================================================================================
# PHẦN 3: ĐIỂM BẮT ĐẦU CHẠY program
# =================================================================================
if __name__ == "__main__":
    try:
        # Chạy bước 1 và nhận về 2 kết quả
        output_df, continue_processing, should_upload ,should_run_filerun = run_parameter_generation()
        
        print("===============================================================================")
        print("=== ステップ➊：パラメータファイル作成プロセスの開始 ===")
        print("===============================================================================")
           
        # Chỉ tiếp tục nếu người dùng không chọn "Kết thúc" ở bước 1
        if continue_processing:
            print("\n=====================================================================")
            print("======ステップ➊ 完了 - ファイルが作成されました。 ======")
            print("=====================================================================")
            
            # CHẠY HÀM RUN_FILERUN VÀ HÀM GOOGLE SHEET MỚI
            run_filerun()
            upload_data_to_google_sheet(output_df) # <-- GỌI HÀM MỚI Ở ĐÂY
            
            # Nếu người dùng đã chọn "Có" upload thì mới chạy phần upload
            if should_upload:
                print("\n\n")
                print("=====================================================================")
                print("=== ステップ➋：クラウドへの自動アップロード開始 ===")
                print("=====================================================================")
                run_auto_uploader()
                # run_filerun()
                print("\n=====================================================================")
                print("====== ステップ➋ 完了 - アップロード成功 ======")
                print("=====================================================================")
            else:
                # run_filerun()
                print("\nアップロードしないことが選択されたため、ステップ➋ をスキップします。.")

    except SystemExit as e:
        logging.warning(f"プログラムが停止しました... {e}")
    except Exception as e:
        logging.error(f"メインプロセス実行中のエラー: {e}")
        raise
    
    print("\nすべて完了しました。ご利用いただきありがとうございます。")
    input("Have a nice day, Press Enter to exit...")