import os
import pandas as pd
import logging
import openpyxl
import config # Import file config
from utils import glob_multi

# --- CÁC HÀM XỬ LÝ CHO LINE ---

def line_setup_directories():
    """Check and create output directory if needed for Line."""
    # Note: Using config variables directly
    for directory in [config.input_dir, config.master_dir]:
        if not os.path.exists(directory):
            logging.error(f"フォルダが見つかりません。: {directory}")
            raise FileNotFoundError(f"フォルダが見つかりません。: {directory}")
    os.makedirs(config.output_dir, exist_ok=True)
    if not os.access(config.output_dir, os.W_OK):
        logging.error(f"Không có quyền ghi vào thư mục: {config.output_dir}")
        raise PermissionError(f"Không có quyền ghi vào thư mục: {config.output_dir}")

def line_read_excel_file(file_path, sheet_name=None, header_row=0, required_cols=None):
    """Read an Excel file, check its existence, and load only required columns for Line."""
    if not os.path.exists(file_path):
        logging.error(f"File '{file_path}' not found. Please check the file name or path.")
        return None, None
    try:
        workbook = openpyxl.load_workbook(file_path)
        logging.info(f"Các sheet in file '{os.path.basename(file_path)}': {workbook.sheetnames}")
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, usecols=required_cols)
        logging.info(f"Tổng số line in sheet '{sheet_name}' (bao gồm tiêu đề): {len(df) + 1}")
        logging.info(f"Các cột in sheet '{sheet_name}': {df.columns.tolist()}")
        return df, workbook.sheetnames
    except Exception as e:
        logging.error(f"error read file '{file_path}': {e}")
        return None, None

def line_read_csv_file(file_path, required_cols=None):
    """Read a CSV file, trying multiple encodings and handling errors for Line."""
    if not os.path.exists(file_path):
        logging.error(f"File '{file_path}' not found. Please check the file name or path.")
        return None
    encodings = ['utf-8-sig', 'utf-16', 'shift-jis', 'latin1']
    for encoding in encodings:
        try:
            df = pd.read_csv(file_path, encoding=encoding, sep='\t', on_bad_lines='warn', dtype=str, usecols=required_cols)
            logging.info(f"Tổng số line in file '{os.path.basename(file_path)}' (bao gồm tiêu đề): {len(df) + 1}")
            logging.info(f"Các cột in file '{os.path.basename(file_path)}': {df.columns.tolist()}")
            logging.info(f"read file '{os.path.basename(file_path)}' sucssecful with encoding: {encoding}")
            return df
        except UnicodeDecodeError as e:
            logging.warning(f"can't read file '{os.path.basename(file_path)}' với encoding {encoding}: {e}")
            continue
        except pd.errors.ParserError as e:
            logging.warning(f"Lỗi phân tích file '{os.path.basename(file_path)}' với encoding {encoding} và sep='\\t': {e}")
            try:
                df = pd.read_csv(file_path, encoding=encoding, sep='\t', on_bad_lines='skip', dtype=str, usecols=required_cols)
                logging.info(f"Tổng số line in file '{os.path.basename(file_path)}' (bao gồm tiêu đề): {len(df) + 1}")
                logging.info(f"Các cột in file '{os.path.basename(file_path)}': {df.columns.tolist()}")
                logging.info(f"read file '{os.path.basename(file_path)}' sucssecful with encoding: {encoding} (bỏ qua line lỗi)")
                return df
            except Exception as e2:
                logging.warning(f"Still can't read file với encoding {encoding} và sep='\\t': {e2}")
                continue
    logging.error(f"can't read file '{file_path}' với bất kỳ encoding nào in {encodings}.")
    return None

def line_check_columns(df, required_cols, source_name):
    """Check if required columns exist in the DataFrame for Line."""
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        logging.error(f"No rule in {source_name}: {missing_cols}")
        raise ValueError(f"Missing columns in {source_name}: {missing_cols}")

def line_check_duplicates(df, column, file_path, log_file):
    """Check for duplicates in a column and log them for Line."""
    duplicates = df[df[column].duplicated(keep=False)]
    if not duplicates.empty:
        logging.warning(f"Có {len(duplicates)} Duplicate values in rule {column}")
        logging.warning(f"Duplicate values: {duplicates[column].unique().tolist()}")
        duplicates.to_csv(log_file, index=False, encoding='utf-8-sig')
        logging.info(f"Duplicate values save in: {log_file}")

def line_process_files():
    """Process files for Line and return output DataFrame."""
    line_setup_directories()
    
    master_file_path = os.path.join(config.master_dir, 'Master.xlsx')
    logging.info("\nRead Master.xlsx:")
    df_master, _ = line_read_excel_file(master_file_path, sheet_name='Sheet1', header_row=0, required_cols=config.LINE_MASTER_REQUIRED_COLS)
    if df_master is None:
        raise SystemExit("Can not continue with read Master.xlsx")
    line_check_columns(df_master, config.LINE_MASTER_REQUIRED_COLS, "sheet 'Sheet1' của Master.xlsx")

    logging.info("\nFind file Master2:")
    # Tìm ở cả Input Folder (user tự đặt thủ công) lẫn Download Folder (vừa tải về)
    master2_files = glob_multi([config.input_dir, config.download_dir], '*campaign-adgroup-ad*.csv')
    if not master2_files:
        raise SystemExit(f"Can't find file .csv  'campaign-adgroup-ad' in  {config.input_dir} hoặc {config.download_dir}")

    df_master2_list = []
    for master2_file in master2_files:
        df = line_read_csv_file(master2_file)
        if df is not None:
            df_master2_list.append(df)
    
    if not df_master2_list:
        raise SystemExit("No 掲載内容.")
    
    df_master2 = pd.concat(df_master2_list, ignore_index=True)
    line_check_columns(df_master2, config.LINE_MASTER2_REQUIRED_COLS, "Master2")
    df_master2['combined_key'] = (df_master2['Campaign name'].astype(str) + '|' + 
                                  df_master2['Ad group name'].astype(str) + '|' + 
                                  df_master2['Ad name'].astype(str))
    if 'Start date' in df_master2.columns:
        df_master2 = df_master2.sort_values('Start date', ascending=False).drop_duplicates('combined_key', keep='first')

    logging.info("\nFind file input:")
    input_files = [f for f in os.listdir(config.input_dir) if '入稿' in f and f.endswith('.xlsm')]
    if not input_files:
        raise SystemExit(f"Can't find file .xlsm  '入稿' in this foder {config.input_dir}")

    df_input_list = []
    for file in input_files:
        input_file = os.path.join(config.input_dir, file)
        df, _ = line_read_excel_file(input_file, sheet_name='AD入稿', header_row=14, required_cols=config.LINE_INPUT_REQUIRED_COLS)
        if df is not None:
            df_input_list.append(df.dropna(subset=config.LINE_INPUT_REQUIRED_COLS, how='all'))

    if not df_input_list:
        raise SystemExit("no file.")

    df_input = pd.concat(df_input_list, ignore_index=True)
    df_input['combined_key'] = (df_input['キャンペーン名'].astype(str) + '|' + 
                                df_input['広告グループ名'].astype(str) + '|' + 
                                df_input['広告名'].astype(str))

    df_output = pd.DataFrame()
    df_output['アカウントID'] = df_input['アカウントID']
    df_output['キャンペーン名'] = df_input['キャンペーン名']
    df_output['広告グループ名'] = df_input['広告グループ名']
    df_output['告名'] = df_input['広告名']
    df_output['パラメ発行済みURL'] = df_input['リンク先URL1']
    df_output['combined_key'] = df_input['combined_key']

    df_output = df_output.merge(df_master[['アカウントID', 'CID']].drop_duplicates(), on='アカウントID', how='left')
    df_output = df_output.merge(
        df_master2[['combined_key', 'Campaign ID', 'Ad group ID', 'Ad ID']].drop_duplicates(), 
        on='combined_key', 
        how='left'
    )
    df_output.drop(columns=['combined_key'], inplace=True, errors='ignore')
    df_output.rename(columns={'Campaign ID': 'キャンペーンID', 'Ad group ID': '広告グループID', 'Ad ID': '広告ID'}, inplace=True)

    for col in config.LINE_OUTPUT_COLUMNS:
        if col not in df_output.columns:
            df_output[col] = ''
    
    for col in df_output.select_dtypes(include=['object']).columns:
        df_output[col] = df_output[col].fillna('')

    df_output['Action'] = config.LINE_ACTION
    df_output['媒体ID'] = config.LINE_MEDIA_ID

    missing_mask = (df_output['キャンペーンID'] == '') | \
                   (df_output['広告グループID'] == '') | \
                   (df_output['広告ID'] == '')
    
    missing_ids_df = df_output[missing_mask]

    if not missing_ids_df.empty:
        logging.warning(f"Có {len(missing_ids_df)} line in output thiếu ít nhất một in các ID (キャンペーンID, 広告グループID, 広告ID)")
        missing_ids_df[['キャンペーン名', '広告グループ名', '告名']].to_csv(config.LINE_MISSING_IDS_FILE, index=False, encoding='utf-8-sig')
        logging.info(f"line thiếu ID được lưu in: {config.LINE_MISSING_IDS_FILE}")

    df_output_final = df_output[~missing_mask]
    logging.info(f"IDが欠損していた {len(missing_ids_df)} 行を最終出力から除外しました。")
    
    df_output_final = df_output_final[config.LINE_OUTPUT_COLUMNS]
    logging.info(f"Số line in file output: {len(df_output_final)}")
    
    return df_output_final